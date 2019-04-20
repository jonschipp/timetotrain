from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from Style import Style
from Utils import Utils


# TODO: Calculate these numbers in dynamically
COLUMN_LENGTH = 7 # The length of each day/slot, determines overall alignment
BEGIN_COLUMN = 2 #  We start in the 2nd column i.e. B for each day/slot
BEGIN_FREQ_ROW = 4 # We start at row 4 for each day/slot
BEGIN_SLOT_ROW = 6 # The first row in a week where the exercise slot begins e.g. [ Exercise 1 ]
NEXT_COLUMN = COLUMN_LENGTH + 2 # Where the next column begins for each day/slot
# TODO: Make these user defineable
# These are updated for each day via update_volume_headers() and reset back to this after each week via reset_volume_headers()
VOLUME_HEADERS = {
    "Sets":    { "ColumnNumber": BEGIN_COLUMN,     "ColumnLetter": get_column_letter(BEGIN_COLUMN)    },
    "Load":    { "ColumnNumber": BEGIN_COLUMN + 1, "ColumnLetter": get_column_letter(BEGIN_COLUMN + 1)},
    "Reps":    { "ColumnNumber": BEGIN_COLUMN + 2, "ColumnLetter": get_column_letter(BEGIN_COLUMN + 2)},
    "RIR":     { "ColumnNumber": BEGIN_COLUMN + 3, "ColumnLetter": get_column_letter(BEGIN_COLUMN + 3)},
    "RPE":     { "ColumnNumber": BEGIN_COLUMN + 4, "ColumnLetter": get_column_letter(BEGIN_COLUMN + 4)},
    "Avg Vel": { "ColumnNumber": BEGIN_COLUMN + 5, "ColumnLetter": get_column_letter(BEGIN_COLUMN + 5)},
    "Int %":   { "ColumnNumber": BEGIN_COLUMN + 6, "ColumnLetter": get_column_letter(BEGIN_COLUMN + 6)},
    "LWL":     { "ColumnNumber": BEGIN_COLUMN + 7, "ColumnLetter": get_column_letter(BEGIN_COLUMN + 7)}
}
VOLUME_LENGTH = len(VOLUME_HEADERS)


class Workout:

  def __init__(self, weeks=8, frequency=3, slots=3, sets=10):
      self.wb = Workbook()
      self.weeks = weeks     # How many weeks for the progrqm
      self.frequency = frequency # How many days per week
      self.slots = slots     # How many slots per day
      self.sets = sets     # How many sets per slot


  def generate_weeks(self, weeks: int) -> list:
      if not weeks:
          # Set default
          weeks = self.weeks

      # Weeks
      for week in range(1, weeks + 1):
         # Create a new sheet for each week
         sheet=f"Week {week}"
         print(f"Writing sheet {sheet}")
         currentSheet = self.wb.create_sheet(title=sheet)

      # Remove default sheet
      del self.wb['Sheet']

      return self.wb.sheetnames


  def generate_frequency(self, frequency: int) -> int:
      if not frequency:
          # Set default
          frequency = self.frequency

      for sheet in self.wb.sheetnames:
          # Get sheet
          # Generate tables for days in sheet
          begin_row = BEGIN_FREQ_ROW
          begin_col = BEGIN_COLUMN
          currentSheet = self.wb[sheet]

          for day in range(1, frequency + 1):
              # Add day header e.g .[ Day 1 ] [ Day 2 ] [ Day 3 ]

              currentCell = Style.generate_header(
                  begin_row, begin_col, COLUMN_LENGTH, currentSheet, heading='Day', value=day
              )

              Style.set_style(
                  currentSheet, currentCell, begin_col,
                  fgColor=Style.Settings.WHITE, bgColor=Style.Settings.LIGHTBLACK,
                  size=42, width=20, font='Helvetica', bold=True
              )

              begin_col += NEXT_COLUMN

          Style.generate_sheet_banner(currentSheet=currentSheet, value=f"{sheet}")

      return frequency


  def generate_slots(self, slots: int, sets: int, frequency: int) -> int:
      # Add exercise slots to chosen frequency e.g.
      if not slots:
          # Set default
          slots = self.slots

      if not frequency:
          # Set default
          frequency = self.frequency

      if not sets:
          # Set default
          sets = self.sets

      for sheet in self.wb.sheetnames:
          # Get sheet
          # Generate tables for days in sheet
          currentSheet = self.wb[sheet]

          # Get beginning column, we adjust as needed in the slot loop
          slot_col = BEGIN_COLUMN

          # These are used to track if for last rows in day
          daily_rpe_row = None
          session_rpe_row = None
          internal_load_row = None

          for day in range(1, frequency + 1):

              # TODO: Determining placement can be done better than this
              slot_rows = {
                  "slot"          : BEGIN_SLOT_ROW,
                  "exercise"      : BEGIN_SLOT_ROW + 1,
                  "programming"   : BEGIN_SLOT_ROW + 2,
                  "target"        : BEGIN_SLOT_ROW + 3,
                  "notes"         : BEGIN_SLOT_ROW + 4,
                  "volume_header" : BEGIN_SLOT_ROW + 5,
                  "volume_input"  : BEGIN_SLOT_ROW + 6,
                  "maxes"         : BEGIN_SLOT_ROW + 6 + sets,
                  "averages"      : BEGIN_SLOT_ROW + 6 + sets + 1,
                  "sums"          : BEGIN_SLOT_ROW + 6 + sets + 2,
                  "volume"        : BEGIN_SLOT_ROW + 6 + sets + 3,
                  "tonnage"       : BEGIN_SLOT_ROW + 6 + sets + 4,
                  "e1rm"          : BEGIN_SLOT_ROW + 6 + sets + 5
              }
              # Used to get the next exercise slot section via its row number
              next_slot=len(slot_rows)+sets
              # Keep track of set range [("C12", "C21"), ..] for Internal Load formula
              set_range=[]

              for slot in range(1, slots + 1):

                  # Add exercise slot header
                  # [    Day 1   ]
                  # [ Exercise 1 ]
                  # [ Exercise 2 ]
                  # [ Exercise 3 ]
                  currentCell = Style.generate_header(
                      slot_rows['slot'], slot_col, COLUMN_LENGTH, currentSheet, heading='Exercise', value=slot
                  )
                  Style.set_style(
                      currentSheet, currentCell, slot_col,
                      fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKGREY,
                      size=32, width=20, font='Helvetica'
                  )
                  slot_rows['slot'] += next_slot

                  # Add exercise header
                  # [    Day 1   ]
                  # [ Exercise 1 ]
                  # [  Exercise  ]
                  # [ Exercise 2 ]
                  # [  Exercise  ]
                  currentCell = Style.generate_header(
                      slot_rows['exercise'], slot_col, COLUMN_LENGTH, currentSheet, heading='Exercise', value=''
                  )
                  Style.set_style(
                      currentSheet, currentCell, slot_col,
                      fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                      size=18, width=20, font='Helvetica', bold=False
                  )
                  slot_rows['exercise'] += next_slot

                  # [       Program      ]
                  # [        Notes       ]
                  # [        Target       ]
                  Style.generate_divide(slot_rows['programming'], slot_col, COLUMN_LENGTH, currentSheet, heading='Program')
                  Style.generate_divide(slot_rows['target'], slot_col, COLUMN_LENGTH, currentSheet, heading='Target')
                  Style.generate_divide(slot_rows['notes'], slot_col, COLUMN_LENGTH, currentSheet, heading='Notes')
                  # TODO: Row height can be set in a better place
                  currentSheet.row_dimensions[slot_rows['programming']].height = 40
                  currentSheet.row_dimensions[slot_rows['target']].height = 40
                  currentSheet.row_dimensions[slot_rows['notes']].height = 40
                  slot_rows['programming'] += next_slot
                  slot_rows['target'] += next_slot
                  slot_rows['notes'] += next_slot

                  # Add set header inputs
                  # [ Sets ] [ Load ] [ Reps ] [ RIR ] [ RPE ] [ Avg Vel ] [ Intensity ]
                  self.generate_volume_header(slot_rows['volume_header'], slot_col, currentSheet)
                  slot_rows['volume_header'] += next_slot

                  # Add set inputs
                  # [ Set 1 ] [ <input> ]
                  # [ Set 2 ] [ <input> ]
                  self.generate_volume_input(slot_rows['volume_input'], slot_col, currentSheet, sets=sets, e1rm_row=slot_rows['e1rm'])
                  set_range.append((f"{VOLUME_HEADERS['Load']['ColumnLetter']}{slot_rows['volume_input']}", f"{VOLUME_HEADERS['Load']['ColumnLetter']}{slot_rows['volume_input']+sets-1}"))

                  # TODO: We should not be be referencing numbers, it's barely readable
                  self.generate_rir_to_rpe(slot_rows['volume_input'], slot_col+4, currentSheet, sets=sets)


                  # Add maxes row
                  # [ Maxes ] [ <formula> ], etc.
                  # Add row for getting the Max (highest number) - for convenience.
                  self.generate_maxes_row(slot_rows['maxes'], slot_col, currentSheet, sets=sets)
                  slot_rows['maxes'] += next_slot

                  # Add averages row
                  # [ Avgs ] [ <formula> ], etc.
                  self.generate_averages_row(slot_rows['averages'], slot_col, currentSheet, sets=sets)
                  slot_rows['averages'] += next_slot

                  # Add averages row
                  # [ Sums ] [ <formula> ], etc.
                  self.generate_sums_row(slot_rows['sums'], slot_col, currentSheet, sets=sets)
                  # Add row for Volume (sets x reps) that reads the Reps sum - for convenience.
                  # Depends on value of slot_rows['sums'] before we increment it
                  Utils.set_formula(
                      currentCell=Style.generate_divide(slot_rows['volume'], slot_col, COLUMN_LENGTH, currentSheet, heading='Volume', style='formula'),
                      formula=f"={VOLUME_HEADERS['Reps']['ColumnLetter']}{slot_rows['sums']}"
                  )
                  slot_rows['sums'] += next_slot

                  Utils.set_formula(
                      currentCell=Style.generate_divide(slot_rows['tonnage'], slot_col, COLUMN_LENGTH, currentSheet, heading='Tonnage', style='formula'),
                      formula=self.generate_tonnage_formula(slot_rows['volume_input'], sets)
                  )
                  slot_rows['tonnage'] += next_slot

                  Utils.set_formula(
                      currentCell=Style.generate_divide(slot_rows['e1rm'], slot_col, COLUMN_LENGTH, currentSheet, heading='E1RM', style='formula'),
                      formula=self.generate_e1rm_formula(slot_rows['volume_input'], sets)
                  )

                  slot_rows['e1rm'] += next_slot
                  slot_rows['volume_input'] += next_slot
                  slot_rows['volume'] += next_slot

              avg_row = slot_rows['averages'] - next_slot
              if not daily_rpe_row:
                  daily_rpe_row = currentSheet.max_row + 1
              if not session_rpe_row:
                  session_rpe_row = currentSheet.max_row + 2
              if not internal_load_row:
                  internal_load_row = currentSheet.max_row + 3
              Utils.set_formula(
                  currentCell=Style.generate_divide(daily_rpe_row, slot_col, COLUMN_LENGTH, currentSheet, heading='Average RPE', style='formula'),
                  formula=f"=IFERROR(AVERAGEIF({VOLUME_HEADERS['RPE']['ColumnLetter']}{avg_row}:{VOLUME_HEADERS['RPE']['ColumnLetter']}{avg_row}, \"<>0\"), \"...\")"
              )
              Utils.set_formula(
                  currentCell=Style.generate_divide(session_rpe_row, slot_col, COLUMN_LENGTH, currentSheet, heading='Session RPE', style='manual'),
                  formula=f""
              )
              Utils.set_formula(
                  currentCell=Style.generate_divide(internal_load_row, slot_col, COLUMN_LENGTH, currentSheet, heading='Internal Load (AU)', style='formula'),
                  formula=self.generate_internal_load_formula(f"{VOLUME_HEADERS['Load']['ColumnLetter']}{session_rpe_row}", set_range)
              )

              # Update starting columns and start writing in column for next day
              self.update_volume_headers()
              slot_col += NEXT_COLUMN

          # Reset volume header back to default position for next week
          self.reset_volume_headers()

      return slots


  def generate_volume_header(self, row: int, col: int, currentSheet: object) -> object:

              for header in VOLUME_HEADERS:
                  currentCell = currentSheet.cell(
                      row=row, column=col, value=f"{header}"
                  )

                  Style.set_style(
                      currentSheet, currentCell, col,
                      fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                      size=12, width=15, font='Helvetica', bold=True
                  )
                  # Set next column
                  col += 1

              return currentCell


  def generate_volume_input(self, row: int, col: int, currentSheet: object, sets: int, **kwargs: dict) -> object:

              for number in range(1, sets + 1):

                  currentCell = currentSheet.cell(
                      row=row, column=col, value=f"Set {number}"
                  )

                  Style.set_style(
                      currentSheet, currentCell, col,
                      fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKGREY,
                      size=12, width=8, font='Helvetica', bold=False
                  )

                  for item in range(1, VOLUME_LENGTH):

                      currentCell = currentSheet.cell(
                          row=row, column=col+item, value=""
                      )

                      # Add intensity calculation based off E1RM
                      if col+item == VOLUME_HEADERS['Int %']['ColumnNumber']:
                          Utils.set_formula(
                              currentCell=currentCell,
                              #=IFERROR(C12/C27, "...")
                              formula=f"=IFERROR({VOLUME_HEADERS['Load']['ColumnLetter']}{row}/{VOLUME_HEADERS['Load']['ColumnLetter']}{kwargs['e1rm_row']}, \"...\")"
                          )
                          currentCell.number_format = '0%'

                      # Add Last Week's Load value
                      if col+item == VOLUME_HEADERS['LWL']['ColumnNumber']:
                          try:
                              self.wb.active = currentSheet
                              # Get last sheet to reference last load
                              last_week = int(self.wb.active.title.split(' ')[1])-1
                              formula=f"=IF(ISBLANK('Week {last_week}'!{VOLUME_HEADERS['Load']['ColumnLetter']}{row}), \"...\", 'Week {last_week}'!{VOLUME_HEADERS['Load']['ColumnLetter']}{row})"
                              # E.g. =IF(ISBLANK('Week 1'!C12), "...", 'Week 1'!C12)
                          except IndexError:
                              formula = "N/A"

                          # The first week which is 0 doesn't have a previous week..skip
                          if last_week > 0:
                              Utils.set_formula(
                                  currentCell=currentCell,
                                  formula=formula
                              )

                      currentCell.alignment = Style.Settings.ALIGNMENT

                  # Set next column
                  row += 1

              return currentCell


  def generate_rir_to_rpe(self, row: int, col: int, currentSheet: object, sets: int) -> object:
              # [ RIR ] to [ RPE ]
              # [  2  ]    [  8  ]
              # e.g. #IF(E12="", "...", ABS(IFERROR(E12−10,"")))

              # The column before contains RPE
              col_rpe_letter = get_column_letter(col-1)

              for input_row in range(row, row + sets):

                  currentCell = currentSheet.cell(
                      row=row, column=col,
                      # Google Sheets: e.g. =IF(ISBLANK(E12),"...", IFERROR(ABS(MINUS(E12, 10)), "N/A"))
                      value=f"=IF(ISBLANK({col_rpe_letter}{input_row}), \"...\", IFERROR(ABS(MINUS({col_rpe_letter}{input_row}, 10)), \"N/A\"))"
                      # Excel: e.g. =IF(ISBLANK(L15), "...", IFERROR(ABS(SUM(L15, -10)), "..."))
                  )

                  currentCell.alignment = Style.Settings.ALIGNMENT

                  row += 1

              return currentCell


  def generate_averages_row(self, row: int, col: int, currentSheet: object, sets: int) -> object:

              currentCell = currentSheet.cell(
                  row=row, column=col, value=f"Averages"
              )

              Style.set_style(
                  currentSheet, currentCell, col,
                  fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                  size=12, width=15, font='Helvetica', bold=True
              )

              col += 1

              # TODO: The hardcoding of position makes it unreadable
              # Get first row of user inputs [ Load ] [ Reps ], etc.
              begin_input_row = row - sets - 1
              # Get last input row [ Load ] [ Reps ], etc.
              end_input_row = row - 2

              for input_row in range(begin_input_row, begin_input_row + COLUMN_LENGTH):

                  col_letter = get_column_letter(col)
                  #print(f"{begin_input_row} + {sets}: {col_letter}")

                  currentCell = currentSheet.cell(
                      row=row, column=col,
                      value=f"=IFERROR(ROUND(AVERAGEIF({col_letter}{begin_input_row}:{col_letter}{end_input_row}, \"<>0\"), 0), \"...\")"
                  )

                  if col == VOLUME_HEADERS['Avg Vel']['ColumnNumber']:
                      currentCell.value = f"=IFERROR(AVERAGEIF({col_letter}{begin_input_row}:{col_letter}{end_input_row}, \"<>0\"), \"...\")"
                  if col == VOLUME_HEADERS['Int %']['ColumnNumber']:
                      currentCell.value = f"=IFERROR(ROUND(AVERAGEIF({col_letter}{begin_input_row}:{col_letter}{end_input_row}, \"<>0\"), 3), \"...\")"

                  Style.set_style(
                      currentSheet, currentCell, col,
                      fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                      size=12, width=8, font='Helvetica', bold=False
                  )

                  currentCell.alignment = Style.Settings.ALIGNMENT

                  if col == VOLUME_HEADERS['Int %']['ColumnNumber']:
                      currentCell.number_format = '0%'

                  # Set next column
                  col += 1

  def generate_sums_row(self, row: int, col: int, currentSheet: object, sets: int) -> object:

              currentCell = currentSheet.cell(
                  row=row, column=col, value=f"Sums"
              )

              Style.set_style(
                  currentSheet, currentCell, col,
                  fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                  size=12, width=15, font='Helvetica', bold=True
              )

              col += 1

              # Get first row of user inputs [ Load ] [ Reps ], etc.
              # TODO: The hardcoding of position makes it unreadable
              begin_input_row = row - sets - 2
              # Get last input row [ Load ] [ Reps ], etc.
              end_input_row = row - 3

              for input_row in range(begin_input_row, begin_input_row + COLUMN_LENGTH):

                  col_letter = get_column_letter(col)

                  currentCell = currentSheet.cell(
                      row=row, column=col,
                      value=f"=IFERROR(IF(SUM({col_letter}{begin_input_row}:{col_letter}{end_input_row})>0, SUM({col_letter}{begin_input_row}:{col_letter}{end_input_row}), \"...\"), \"N/A\")"
                  )

                  Style.set_style(
                      currentSheet, currentCell, col,
                      fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                      size=12, width=8, font='Helvetica', bold=False
                  )

                  currentCell.alignment = Style.Settings.ALIGNMENT

                  if col == VOLUME_HEADERS['Int %']['ColumnNumber']:
                      currentCell.number_format = '0%'

                  # Set next column
                  col += 1


  def generate_maxes_row(self, row: int, col: int, currentSheet: object, sets: int) -> object:

              currentCell = currentSheet.cell(
                  row=row, column=col, value=f"Maxes"
              )

              Style.set_style(
                  currentSheet, currentCell, col,
                  fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                  size=12, width=15, font='Helvetica', bold=True
              )

              col += 1

              # Get first row of user inputs [ Load ] [ Reps ], etc.
              begin_input_row = row - sets
              # Get last input row [ Load ] [ Reps ], etc.
              end_input_row = row - 1

              for input_row in range(begin_input_row, begin_input_row + COLUMN_LENGTH):

                  col_letter = get_column_letter(col)

                  currentCell = currentSheet.cell(
                      row=row, column=col,
                      value=f"=IFERROR(MAX({col_letter}{begin_input_row}:{col_letter}{end_input_row}), \"...\")"
                      #E.g. IFERROR(MAX(C12:C16), "...")
                  )

                  Style.set_style(
                      currentSheet, currentCell, col,
                      fgColor=Style.Settings.WHITE, bgColor=Style.Settings.DARKRED,
                      size=12, width=8, font='Helvetica', bold=False
                  )

                  currentCell.alignment = Style.Settings.ALIGNMENT

                  if col == VOLUME_HEADERS['Int %']['ColumnNumber']:
                      currentCell.number_format = '0%'

                  # Set next column
                  col += 1

  def generate_tonnage_formula(self, row, sets) -> str:
      #=SUM(PRODUCT(C34:C34),PRODUCT(C35:C35),PRODUCT(C36:C36)...)
      l = []

      first_row = row
      last_row = row + sets - 1
      for r in range(row, row + sets):
          l.append(f"{VOLUME_HEADERS['Load']['ColumnLetter']}{r}:{VOLUME_HEADERS['Reps']['ColumnLetter']}{r}")

      formula = '{}, {}{}), {})'.format(
          f"=IF(COUNT({VOLUME_HEADERS['Load']['ColumnLetter']}{first_row}:{VOLUME_HEADERS['Load']['ColumnLetter']}{last_row})>0",
          f"SUM(", "".join('PRODUCT({}), '.format(i) for i in l).rstrip(','),
          f"\"...\""
      )
      return formula

  def generate_e1rm_formula(self, row, sets) -> str:
      # Epley equation W * (1 + r/30)
      # $ echo "315 * (1 + 5/30)" | bc -l
      # 367.49999999999999999790
      # E.g. =MAX(C12:C21)*(1+VLOOKUP(MAX(C12:C21),C:D,2, FALSE)/30)

      first_row = row
      last_row = row + sets - 1

      # =IFERROR(PRODUCT(MAX(C12:C21), SUM(1, DIVIDE(VLOOKUP(MAX(C12:C21), C12:D21, 2,FALSE), 30))), "...")
      formula = '{}{}{}'.format(
          f"=IFERROR(PRODUCT(MAX({VOLUME_HEADERS['Load']['ColumnLetter']}{first_row}:{VOLUME_HEADERS['Load']['ColumnLetter']}{last_row}), ",
          f"SUM(1, DIVIDE(VLOOKUP(MAX({VOLUME_HEADERS['Load']['ColumnLetter']}{first_row}:{VOLUME_HEADERS['Load']['ColumnLetter']}{last_row}), ",
          f"{VOLUME_HEADERS['Load']['ColumnLetter']}{first_row}:{VOLUME_HEADERS['Reps']['ColumnLetter']}{last_row}, 2, FALSE), 30))), \"...\")"
      )

      return formula

  def generate_internal_load_formula(self, session_cell, set_range) -> str:
      # =IF(ISBLANK(C52), "...", PRODUCT(C52, SUM(COUNTIF(C12:C21, ">0"), COUNTIF(C35:C44, ">0"))))

      formula = '=IF(ISBLANK({}), {}, PRODUCT({}, SUM({})))'.format(
          f"{session_cell}",
          f"\"...\"",
          f"{session_cell}",
          f"".join('COUNTIF({}:{}, ">0"), '.format(j,k) for j,k in set_range).rstrip(", '"),
      )
      return formula

  def update_volume_headers(self) -> None:
      for k in VOLUME_HEADERS.keys():
          VOLUME_HEADERS[k]['ColumnNumber'] += 9
          VOLUME_HEADERS[k]['ColumnLetter'] = get_column_letter(VOLUME_HEADERS[k]['ColumnNumber'])
      return None


  def reset_volume_headers(self) -> None:
      for k, n in  zip(VOLUME_HEADERS.keys(), range(0, VOLUME_LENGTH)):
          VOLUME_HEADERS[k]['ColumnNumber'] = BEGIN_COLUMN + n
          VOLUME_HEADERS[k]['ColumnLetter'] = get_column_letter(BEGIN_COLUMN + n)


  def test(self, msg: str) -> str:
      if not msg:
          msg = 'Test'
      return msg
